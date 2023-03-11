import openpyxl
import os
import pathlib

import xml.etree.ElementTree as ET
from xml.dom import minidom

# settings
excel_path = 'C:/Temp/PPC_HLM 2023.xlsx'
export_path = 'C:/Temp/export.xml'

class importFsmExport:
    def __init__(self, xml_path) -> None:

        self.xml_path = pathlib.Path(xml_path)
        xml = None
        try:
            xml = ET.parse(xml_path)
        except  Exception as e:
            print('Unable to parse xml file.')
            raise e
        self.xml = xml
        self.output_files = {} # dict of ConvertedOutputType -> list of file names (list(Path))

    def findName(self, givenName: str, familyName: str) -> ET.Element:
        #participants = self.xml.findall(f".//Participant[@GivenName='{givenName}'")
        participants = self.xml.findall(f".//Participant")
        participants_filtered = [e for e in participants if e.attrib['GivenName'] == givenName and e.attrib['FamilyName'] == familyName]
        if len(participants_filtered) == 0:
            print(f"Not found in export: {givenName} {familyName}")
            return None
        if len(participants_filtered) > 1:
            print(f"Found more than once in export: {givenName} {familyName}")
        return participants_filtered[0]

    def setPpcs(self, ppcs: { "short": [str], "free": [str] }, participant: ET.Element) -> None:
        def createPpcElement(e: str, i: int, short: bool):
            return ET.Element('EventEntry', { "Type": "ER_EXTENDED",  "Code": "ELEMENT_CODE_SHORT" if short else "ELEMENT_CODE_FREE", "Pos": str(i),  "Value": e})
        registeredEvent = participant.find("./Discipline/RegisteredEvent")
        if registeredEvent:
            print('Written: ' + participant.attrib['GivenName'])
            for subelement in list(registeredEvent):
                registeredEvent.remove(subelement)
            ppcElements = [createPpcElement(e, i+1, short = True) for i,e in enumerate(ppcs['short'])] + [createPpcElement(e, i+1, short = False) for i,e in enumerate(ppcs['free'])]
            registeredEvent.extend(ppcElements)
        else:
            print('NOT Written: ' + participant.attrib['GivenName'])

    def write(self):
        xmlstr = minidom.parseString(ET.tostring(self.xml.getroot(), xml_declaration= True)).toprettyxml(indent="  ")
        with open("c:\\temp\\output.xml", "w", encoding="utf-8") as f:
            f.write(xmlstr)

class ImportPpc:
    def __init__(self, xlsx_path) -> None:

        self.xlsx_path = pathlib.Path(xlsx_path)
        wb = None
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except  Exception as e:
            print('Unable to parse xlsx file.')
            raise e

        if not wb:
            raise Exception('Unable to load workbook from "%s".' % xlsx_path)

        if len(wb.worksheets) < 1:
            raise Exception('There is no worksheet within the xlsx file "%s"' % xlsx_path)

        self.worksheet = wb.worksheets[0] #[wb.sheetnames[0]]
        self.output_files = {} # dict of ConvertedOutputType -> list of file names (list(Path))

    def convert(self) -> [((str, str), {'short': [str], 'free': [str]})]:
        ret = [];
        iter_ = self.worksheet.iter_rows()
        for row in iter_:
            if row[6].value != None:
                name = (row[6].value, row[7].value)
                ppcs_short = []
                ppcs_free = []
                for row2 in iter_:
                    finish = True
                    if row2[5].value != None:
                        ppcs_free.append(row2[5].value)
                        finish = False
                    if row2[4].value != None:
                        ppcs_short.append(row2[4].value)
                        finish = False
                    if finish:
                        break
                ret.append((name, {'short': ppcs_short, 'free': ppcs_free}))
                if len(ppcs_free) == 0:
                    print(f"{name[0]} has an empty list for free-style")
        return ret

# main script code
if __name__ == '__main__':
    if os.path.isfile(excel_path):
        # try:
            deu_xlsx = ImportPpc(excel_path)
            ppc_list = deu_xlsx.convert()
        # except Exception as e:
        #    print('Failed')
        #    print(e)

    if os.path.isfile(export_path):
        # try:
            xml = importFsmExport(export_path)
            for (name, ppcs) in ppc_list:
                nameEl = xml.findName(name[0], name[1])
                if nameEl:
                   #print(ppcs)
                   xml.setPpcs(ppcs, nameEl)
            xml.write()

        # except Exception as e:
        #    print('Failed')
        #    print(e)
