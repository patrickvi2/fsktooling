import csv
from datetime import datetime, date
from enum import Enum
import logging
import openpyxl
import os
import pathlib
from typing import List

# settings
input_path = 'BJM22Test/Meldungen/' # file -> convert file only; folder -> convert all in folder

class DEUFormType(Enum):
    EVENT = ('001v3') # Eventmeldung
    TEAMEVENT = ('002v1') # Teammeldung
    # SPORTPASS = ('003v1') # Sportpass
    # ORGANIZATION = ('004', 'v3') # Organisationsmeldung

class DEUEventType(Enum):
    WETTBEWERB = ('Wettbewerb')
    KLASSENLAUFEN = ('Klassenlaufen')
    SPORTLER_LEHRGANG = ('Sportler Lehrgang')
    OFFIZIELLEN_LEHRGANG = ('Offiziellen Lehrgang')
    TRAINER_LEHRGANG = ('Trainer Lehrgang')
    SONSTIGE = ('Sonstige')

class ConvertedOutputType(Enum):
    EVENT_INFO = 0
    EVENT_CATERGORIES = 1
    EVENT_PERSONS = 2


class DEUMeldeformularXLSX:
    def __init__(self, xlsx_path, create_categories=True, create_persons=True, overwrite_output_files=True) -> None:

        self.xlsx_path = pathlib.Path(xlsx_path)
        self.create_categories = create_categories
        self.create_persons = create_persons
        self.overwrite = overwrite_output_files
        wb = None
        try:
            wb = openpyxl.load_workbook(xlsx_path)
        except  Exception as e:
            print('Unable to parse xlsx file.')
            raise e

        if not wb:
            raise Exception('Unable to load workbook from "%s".' % xlsx_path)

        if len(wb.worksheets) < 1:
            raise Exception('There is no worksheet within the xlsx file "%s"' % xlsx_path)

        self.worksheet = wb.worksheets[0] #[wb.sheetnames[0]]
        self.output_files = {} # dict of ConvertedOutputType -> list of file names (list(Path))

    @property
    def form_type(self):
        try:
            return DEUFormType(self.worksheet['B1'].value)
        except:
            return None

    @property
    def event_name(self) -> str:
        return str(self.worksheet['F2'].value)

    @property
    def event_organizer(self) -> str:
        return str(self.worksheet['F4'].value)


    @property
    def event_place(self) -> str:
        return str(self.worksheet['F9'].value)

    @property
    def event_type(self) -> DEUEventType:
        try:
            return DEUEventType(self.worksheet['C9'].value)
        except:
            return None

    @staticmethod
    def cell_to_date(cell) -> date:
        if type(cell.value) == datetime:
            return cell.value.date()
        elif type(cell.value) == str:
            return datetime.fromisoformat(cell.value).date()

    @property
    def event_start_date(self) -> date:
        try:
            return self.cell_to_date(self.worksheet['C14'])
        except:
            # default to 1. January of current year
            d = date.today()
            return date(d.year, 1, 1)
    
    @property
    def event_end_date(self) -> date:
        try:
            return self.cell_to_date(self.worksheet['F14'])
        except:
            return self.event_start_date()

    def create_csv_from_table_range(self, min_row, max_row, output_csv_file_path):
        header = []
        csv_content_list = []
        for row in self.worksheet.iter_rows(min_row=min_row, max_row=max_row):
            if not header:
                for cell in row:
                    if cell.value:
                        header.append(cell.value)
                continue
            else:
                if not row[0].value:
                    continue
                csv_dict = {}
                for header_name, cell in zip(header, row):
                    csv_dict[header_name] = cell.value if cell.value else ''
                csv_content_list.append(csv_dict)

        self.write_csv(csv_content_list, output_csv_file_path)

    def write_csv(self, csv_data: dict, path: str) -> None:
        if len(csv_data) == 0:
            print('Warning: CSV data is empty for "%s"' % path)
            return

        if not self.overwrite and os.path.exists(path):
            print('Warning: File "%s" exists already. No csv file created. Set overwrite=True to overwrite a existing csv file.' % path)
            return

        with open(path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
            writer.writeheader()
            writer.writerows(csv_data)

    def get_output_files(self, file_type: ConvertedOutputType) -> List[pathlib.Path]:
        if file_type not in self.output_files:
            return []
        return self.output_files[file_type]

    def add_output_file(self, file_type: ConvertedOutputType, path: pathlib.Path) -> None:
        if file_type not in self.output_files:
            self.output_files[file_type] = []
        self.output_files[file_type].append(path)

    def convert(self):
        if self.form_type is None:
            print('Warning: unable to verify document type and version.')
            print('Still trying by assuming an event registration form...')

        if self.form_type is DEUFormType.EVENT or self.form_type is None:
            self.convert_to_csv()
            # TODO: generate Klassenlaufen scripts
        # TODO: support more types (e.g. sys team registration)

    def convert_to_csv(self):
        base_path = self.xlsx_path.parent / 'csv'
        base_path.mkdir(exist_ok=True)
        output_competition_csv_file_path = base_path / ( self.xlsx_path.stem + '_deu_competition.csv')
        output_categories_csv_file_path = base_path / ( self.xlsx_path.stem + '_deu_categories.csv')
        output_persons_csv_file_path = base_path / ( self.xlsx_path.stem + '_deu_athletes.csv')

        competition_data = [{
            'Name' : self.event_name,
            'Veranstalter' : self.event_organizer,
            'Ort' : self.event_place,
            'Start Datum' : self.event_start_date,
            'End Datum' : self.event_end_date
        }, ]
        self.write_csv(competition_data, output_competition_csv_file_path)
        self.add_output_file(ConvertedOutputType.EVENT_INFO, output_competition_csv_file_path)

        start_row_categories = 0
        end_row_categories = 0
        start_row_persons = 0
        for row in self.worksheet.iter_rows(min_row=1, max_col=2, max_row=2000):
            if row[1].value == 'Team ID':
                start_row_persons = row[1].row
                break
            if row[1].value == 'Disziplin':
                start_row_categories = row[0].row
            elif row[0].value == 'Hinweise:':
                end_row_categories = row[1].row - 1

        if self.create_categories:
            if start_row_categories and end_row_categories and start_row_categories != end_row_categories:
                self.create_csv_from_table_range(start_row_categories, end_row_categories, output_categories_csv_file_path)
                self.add_output_file(ConvertedOutputType.EVENT_CATERGORIES, output_categories_csv_file_path)
            else:
                print('Categories not found in "%s"' % self.xlsx_path)
        if self.create_persons:
            if start_row_persons:
                self.create_csv_from_table_range(start_row_persons, None, output_persons_csv_file_path)
                self.add_output_file(ConvertedOutputType.EVENT_PERSONS, output_persons_csv_file_path)
            else:
                print('Persons not found in "%s"' % self.xlsx_path)

def convert_meldeformular_in_directory(input_directory, create_categories=True):
    for dir_elem in os.listdir(input_directory):
        input_file_path = os.path.join(input_directory, dir_elem)
        if not os.path.isfile(input_file_path):
            continue # not a file

        file_name_and_ext = os.path.splitext(dir_elem)
        if len(file_name_and_ext) < 2:
            continue # no extension

        if file_name_and_ext[1].lower() != '.xlsx':
            continue # not a xlsx extension

        try:
            deu_xlsx = DEUMeldeformularXLSX(input_file_path, create_categories)
            deu_xlsx.convert()
            # create_categories = False # uncomment this, if only the categories of the first file should be used
        except Exception as e:
            print('Error: Failed to convert %s' % input_file_path)
            print(e)


# main script code
if __name__ == '__main__':
    if os.path.isfile(input_path):
        try:
            deu_xlsx = DEUMeldeformularXLSX(input_path)
            deu_xlsx.convert()
        except Exception as e:
            print('Failed')
            print(e)
    elif os.path.isdir(input_path):
        create_categories = True # only for first xlsx
        convert_meldeformular_in_directory(input_path, create_categories)

    # deu_xlsx = DEUMeldeformularXLSX(input_path)
    # print(deu_xlsx.event_name)
    # print(deu_xlsx.event_start_date)
    # print(type(deu_xlsx.event_start_date))

else:
    print = logging.info