import os
import csv
import openpyxl

# settings

input_path = 'OBM22/csv/OBM' # file -> convert file only; folder -> convert all in folder


# functions

def create_csv_from_table_range(worksheet, min_row, max_row, output_csv_file_path, overwrite):
    header = []
    csv_content_list = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row):
        if not header:
            for cell in row:
                if cell.value:
                    header.append(cell.value)
            continue
        else:
            if not row[0].value:
                continue
            csv_dict = {}
            for header_name, cell in zip(header, row):
                csv_dict[header_name] = cell.value if cell.value else ''
            csv_content_list.append(csv_dict)

    if overwrite or not os.path.exists(output_csv_file_path):
        with open(output_csv_file_path, 'w') as f:
            writer = csv.DictWriter(f, fieldnames=header)
            writer.writeheader()
            writer.writerows(csv_content_list)

def convert_meldeformular(input_file_path, create_categories, create_athletes, overwrite_output_files=True):

    output_categories_csv_file_path = os.path.splitext(input_file_path)[0] + '_deu_categories.csv'
    output_participant_csv_file_path = os.path.splitext(input_file_path)[0] + '_deu_athletes.csv'

    wb = None
    try:
        wb = openpyxl.load_workbook(input_file_path)
    except:
        print('Unable to open xlsx file.')

    if not wb:
        print('No workbook loaded.')
        exit(1)

    if len(wb.sheetnames) < 1:
        print('There is no worksheet within the xlsx file.')
        exit(2)

    ws = wb[wb.sheetnames[0]]

    start_row_events = 0
    end_row_events = 0
    start_row_participants = 0
    for row in ws.iter_rows(min_row=1, max_col=2, max_row=2000):
        if row[1].value == 'Team ID':
            start_row_participants = row[1].row
            break
        if row[1].value == 'Disziplin':
            start_row_events = row[0].row
        elif row[0].value == 'Hinweise:':
            end_row_events = row[1].row - 1

    if create_categories:
        if start_row_events and end_row_events and start_row_events != end_row_events:
            create_csv_from_table_range(ws, start_row_events, end_row_events, output_categories_csv_file_path, overwrite_output_files)
        else:
            print('Events not found in "%s"' % input_file_path)
    if create_athletes:
        if start_row_participants:
            create_csv_from_table_range(ws, start_row_participants, None, output_participant_csv_file_path, overwrite_output_files)
        else:
            print('Participants not found in "%s"' % input_file_path)

def convert_meldeformular_in_directory(input_directory, create_categories=True):
    for dir_elem in os.listdir(input_directory):
        input_file_path = os.path.join(input_directory, dir_elem)
        if not os.path.isfile(input_file_path):
            continue # not a file

        file_name_and_ext = os.path.splitext(dir_elem)
        if len(file_name_and_ext) < 2:
            continue # no extension

        if file_name_and_ext[1].lower() != '.xlsx':
            continue # not a xlsx extension

        convert_meldeformular(input_file_path, create_categories, True)
        create_categories = False

# main script code
if os.path.isfile(input_path):
    convert_meldeformular(input_path, False, True)
elif os.path.isdir(input_path):
    create_categories = True # only for first xlsx
    convert_meldeformular_in_directory(input_path)
    
